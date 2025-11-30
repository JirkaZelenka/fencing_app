"""
Django management command to create an example Excel template file.

Usage:
    python manage.py create_excel_template output_file.xlsx
"""

from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta


class Command(BaseCommand):
    help = 'Create an example Excel template file for importing data'

    def add_arguments(self, parser):
        parser.add_argument(
            'output_file',
            type=str,
            help='Path to the output Excel file (.xlsx)'
        )

    def handle(self, *args, **options):
        output_file = options['output_file']
        
        if not output_file.endswith('.xlsx'):
            output_file += '.xlsx'
        
        self.stdout.write(f'Creating Excel template: {output_file}')
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Users sheet
        self._create_users_sheet(wb)
        
        # Create Events sheet
        self._create_events_sheet(wb)
        
        # Create Participations sheet
        self._create_participations_sheet(wb)
        
        wb.save(output_file)
        self.stdout.write(self.style.SUCCESS(f'Template created successfully: {output_file}'))
        self.stdout.write('\nYou can now fill in the data and use:')
        self.stdout.write(f'  python manage.py import_from_excel {output_file}')

    def _create_users_sheet(self, wb):
        """Create Users sheet with headers and example data."""
        ws = wb.create_sheet('Users')
        
        # Headers
        headers = ['username', 'email', 'first_name', 'last_name', 'club_name', 'phone', 'gender', 'birth_year']
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Example data rows
        examples = [
            ['john.doe', 'john.doe@example.com', 'John', 'Doe', 'Fencing Club Prague', '+420123456789', 'M', '2000'],
            ['jane.smith', 'jane.smith@example.com', 'Jane', 'Smith', 'Fencing Club Prague', '+420987654321', 'Z', '2001'],
            ['', 'bob.wilson@example.com', 'Bob', 'Wilson', 'Fencing Club Brno', '', 'M', '1999'],
            ['alice.brown', '', 'Alice', 'Brown', 'Fencing Club Ostrava', '+420555123456', 'Z', '2002'],
        ]
        
        for row in examples:
            ws.append(row)
        
        # Adjust column widths
        column_widths = [15, 25, 12, 12, 20, 18, 8, 12]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
        
        # Add note
        ws.append([])
        ws.append(['Note: At least one of username, email, or both first_name and last_name must be provided.'])

    def _create_events_sheet(self, wb):
        """Create Events sheet with headers and example data."""
        ws = wb.create_sheet('Events')
        
        # Headers
        headers = ['title', 'date', 'location', 'description', 'event_type', 'gender', 'participants_count', 'external_link']
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Example data rows
        today = datetime.now()
        examples = [
            [
                'Czech Fencing Championship 2024',
                (today + timedelta(days=30)).strftime('%Y-%m-%d'),
                'Prague Sports Hall',
                'Annual national championship tournament',
                'tournament',
                'V',
                '40',
                'https://www.czechfencing.cz/tournament/2024'
            ],
            [
                'Regional Tournament - Spring 2024',
                (today + timedelta(days=60)).strftime('%Y-%m-%d'),
                'Brno Fencing Center',
                'Regional qualification tournament',
                'tournament',
                'M',
                '25',
                ''
            ],
            [
                'Charity Fencing Event',
                (today + timedelta(days=90)).strftime('%Y-%m-%d'),
                'Ostrava Arena',
                'Humanitarian tournament for charity',
                'humanitarian',
                'V',
                '30',
                ''
            ],
        ]
        
        for row in examples:
            ws.append(row)
        
        # Adjust column widths
        column_widths = [30, 15, 20, 30, 15, 8, 18, 40]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
        
        # Add note
        ws.append([])
        ws.append(['Note: Date format: YYYY-MM-DD (date only, no time)'])
        ws.append(['Event types: tournament, humanitarian, other'])
        ws.append(['Gender: M (male), Z (female), V (all)'])
        ws.append(['participants_count: Required for tournament and humanitarian events, must be > 0'])

    def _create_participations_sheet(self, wb):
        """Create Participations sheet with headers and example data."""
        ws = wb.create_sheet('Participations')
        
        # Headers
        headers = ['fencer_identifier', 'event_title', 'date', 'position', 'wins', 'losses', 'touches_scored', 'touches_received', 'points']
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Example data rows
        today = datetime.now()
        examples = [
            [
                'john.doe',
                'Czech Fencing Championship 2024',
                (today + timedelta(days=30)).strftime('%Y-%m-%d'),
                '5',
                '8',
                '3',
                '45',
                '32',
                '12.5'
            ],
            [
                'jane.smith',
                'Czech Fencing Championship 2024',
                (today + timedelta(days=30)).strftime('%Y-%m-%d'),
                '2',
                '10',
                '1',
                '52',
                '28',
                '25.0'
            ],
            [
                'bob.wilson@example.com',
                'Regional Tournament - Spring 2024',
                (today + timedelta(days=60)).strftime('%Y-%m-%d'),
                '1',
                '12',
                '0',
                '60',
                '25',
                '30.0'
            ],
            [
                'Alice Brown',
                'Charity Fencing Event',
                (today + timedelta(days=90)).strftime('%Y-%m-%d'),
                '3',
                '7',
                '2',
                '38',
                '30',
                '15.5'
            ],
        ]
        
        for row in examples:
            ws.append(row)
        
        # Adjust column widths
        column_widths = [20, 30, 15, 10, 8, 8, 15, 15, 10]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
        
        # Add note
        ws.append([])
        ws.append(['Note: fencer_identifier must match username, email, or "first_name last_name" from Users sheet'])
        ws.append(['event_title must match exactly with title from Events sheet'])
        ws.append(['If multiple events have same title, include date to specify which one'])


