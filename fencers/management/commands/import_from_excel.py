"""
Django management command to import Users, Events, and Event Participations from Excel.

This command reads data from an Excel file with multiple sheets and populates the database
in the correct order: Users -> FencerProfiles -> Events -> EventParticipations.

Usage:
    python manage.py import_from_excel path/to/file.xlsx
"""

import os
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from django.db import transaction
from openpyxl import load_workbook
from datetime import datetime

from fencers.models import FencerProfile, Club, Event, EventParticipation


class Command(BaseCommand):
    help = 'Import Users, Events, and Event Participations from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Path to the Excel file (.xlsx)'
        )
        parser.add_argument(
            '--create-users',
            action='store_true',
            help='Create User accounts if they do not exist (default: False)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Perform a dry run without saving to database',
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        create_users = options['create_users']
        dry_run = options['dry_run']

        if not os.path.exists(excel_file):
            raise CommandError(f'Excel file not found: {excel_file}')

        self.stdout.write(self.style.SUCCESS(f'Loading Excel file: {excel_file}'))
        
        try:
            workbook = load_workbook(excel_file, data_only=True)
        except Exception as e:
            raise CommandError(f'Error loading Excel file: {e}')

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN MODE - No data will be saved'))

        try:
            with transaction.atomic():
                # Step 1: Import Users and FencerProfiles
                self.stdout.write(self.style.SUCCESS('\n=== Step 1: Importing Users and FencerProfiles ==='))
                users_data = self._read_users_sheet(workbook)
                user_mapping = self._import_users(users_data, create_users, dry_run)
                fencer_mapping = self._import_fencer_profiles(users_data, user_mapping, dry_run)

                # Step 2: Import Events
                self.stdout.write(self.style.SUCCESS('\n=== Step 2: Importing Events ==='))
                events_data = self._read_events_sheet(workbook)
                event_mapping = self._import_events(events_data, dry_run)

                # Step 3: Import Event Participations
                self.stdout.write(self.style.SUCCESS('\n=== Step 3: Importing Event Participations ==='))
                participations_data = self._read_participations_sheet(workbook)
                self._import_participations(participations_data, fencer_mapping, event_mapping, dry_run)

                if dry_run:
                    self.stdout.write(self.style.WARNING('\n=== DRY RUN COMPLETE - No data was saved ==='))
                    raise transaction.TransactionManagementError('Dry run - rolling back transaction')
                else:
                    self.stdout.write(self.style.SUCCESS('\n=== Import completed successfully! ==='))

        except transaction.TransactionManagementError:
            # Expected for dry run
            pass
        except Exception as e:
            raise CommandError(f'Error during import: {e}')

    def _read_users_sheet(self, workbook):
        """Read Users sheet from workbook."""
        if 'Users' not in workbook.sheetnames:
            self.stdout.write(self.style.WARNING('No "Users" sheet found, skipping user import'))
            return []

        sheet = workbook['Users']
        users_data = []
        
        # Read header row
        headers = [cell.value for cell in sheet[1]]
        expected_headers = ['username', 'email', 'first_name', 'last_name', 'club_name', 'phone', 'gender']
        
        # Validate headers
        if not all(header in headers for header in expected_headers):
            self.stdout.write(self.style.WARNING(
                f'Expected headers: {expected_headers}, found: {headers}'
            ))

        # Read data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if not any(cell.value for cell in row):  # Skip empty rows
                continue
            
            row_data = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    row_data[header] = row[col_idx].value
            
            # Only add if we have at least username or name
            if row_data.get('username') or (row_data.get('first_name') and row_data.get('last_name')):
                users_data.append(row_data)
            else:
                self.stdout.write(self.style.WARNING(
                    f'Row {row_idx}: Skipping row without username or name'
                ))

        self.stdout.write(f'Found {len(users_data)} user(s) to import')
        return users_data

    def _read_events_sheet(self, workbook):
        """Read Events sheet from workbook."""
        if 'Events' not in workbook.sheetnames:
            raise CommandError('"Events" sheet is required but not found')

        sheet = workbook['Events']
        events_data = []
        
        # Read header row
        headers = [cell.value for cell in sheet[1]]
        expected_headers = ['title', 'date', 'location', 'description', 'event_type', 'gender', 'participants_count', 'external_link']
        
        # Read data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if not any(cell.value for cell in row):  # Skip empty rows
                continue
            
            row_data = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    row_data[header] = row[col_idx].value
            
            # Require at least title and date
            if row_data.get('title') and row_data.get('date'):
                events_data.append(row_data)
            else:
                self.stdout.write(self.style.WARNING(
                    f'Row {row_idx}: Skipping event without title or date'
                ))

        self.stdout.write(f'Found {len(events_data)} event(s) to import')
        return events_data

    def _read_participations_sheet(self, workbook):
        """Read Participations sheet from workbook."""
        if 'Participations' not in workbook.sheetnames:
            self.stdout.write(self.style.WARNING('No "Participations" sheet found, skipping participations import'))
            return []

        sheet = workbook['Participations']
        participations_data = []
        
        # Read header row
        headers = [cell.value for cell in sheet[1]]
        expected_headers = ['fencer_identifier', 'event_title', 'date', 'position', 'wins', 'losses', 'touches_scored', 'touches_received']
        
        # Read data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if not any(cell.value for cell in row):  # Skip empty rows
                continue
            
            row_data = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    row_data[header] = row[col_idx].value
            
            # Require fencer identifier and event identifier
            if row_data.get('fencer_identifier') and (row_data.get('event_title') or row_data.get('date')):
                participations_data.append(row_data)
            else:
                self.stdout.write(self.style.WARNING(
                    f'Row {row_idx}: Skipping participation without fencer or event identifier'
                ))

        self.stdout.write(f'Found {len(participations_data)} participation(s) to import')
        return participations_data

    def _import_users(self, users_data, create_users, dry_run):
        """Import users and return mapping of username/email to User object."""
        user_mapping = {}
        
        for user_data in users_data:
            username = user_data.get('username')
            email = user_data.get('email', '')
            first_name = user_data.get('first_name', '')
            last_name = user_data.get('last_name', '')
            
            # Try to find existing user
            user = None
            if username:
                try:
                    user = User.objects.get(username=username)
                    self.stdout.write(f'  Found existing user: {username}')
                except User.DoesNotExist:
                    pass
            
            if not user and email:
                try:
                    user = User.objects.get(email=email)
                    self.stdout.write(f'  Found existing user by email: {email}')
                except User.DoesNotExist:
                    pass
            
            # Create user if not found and create_users is True
            if not user:
                if create_users:
                    if not username:
                        # Generate username from name
                        username = f"{first_name.lower()}.{last_name.lower()}".replace(' ', '')
                        # Ensure uniqueness
                        base_username = username
                        counter = 1
                        while User.objects.filter(username=username).exists():
                            username = f"{base_username}{counter}"
                            counter += 1
                    
                    if not email:
                        email = f"{username}@example.com"
                    
                    if not dry_run:
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            first_name=first_name,
                            last_name=last_name
                        )
                        self.stdout.write(self.style.SUCCESS(f'  Created user: {username}'))
                    else:
                        self.stdout.write(f'  [DRY RUN] Would create user: {username}')
                        user = type('User', (), {'id': None, 'username': username})()  # Mock user
                else:
                    self.stdout.write(self.style.WARNING(
                        f'  User not found: {username or email}. Use --create-users to create new users.'
                    ))
                    continue
            
            # Store mapping
            identifier = username or email or f"{first_name} {last_name}".strip()
            user_mapping[identifier] = user
        
        return user_mapping

    def _import_fencer_profiles(self, users_data, user_mapping, dry_run):
        """Import fencer profiles and return mapping of identifier to FencerProfile."""
        fencer_mapping = {}
        
        for user_data in users_data:
            username = user_data.get('username')
            email = user_data.get('email', '')
            first_name = user_data.get('first_name', '')
            last_name = user_data.get('last_name', '')
            club_name = user_data.get('club_name', '')
            phone = user_data.get('phone', '')
            gender = user_data.get('gender', '')
            
            identifier = username or email or f"{first_name} {last_name}".strip()
            user = user_mapping.get(identifier)
            
            if not user:
                continue
            
            # Get or create club
            club = None
            if club_name:
                if not dry_run:
                    club, created = Club.objects.get_or_create(name=club_name)
                    if created:
                        self.stdout.write(f'  Created club: {club_name}')
                else:
                    self.stdout.write(f'  [DRY RUN] Would get/create club: {club_name}')
                    club = type('Club', (), {'id': None, 'name': club_name})()  # Mock club
            
            # Get or create fencer profile
            if not dry_run:
                if hasattr(user, 'id') and user.id:  # Real user object
                    fencer_profile, created = FencerProfile.objects.get_or_create(
                        user=user,
                        defaults={
                            'club': club,
                            'phone': phone or '',
                            'gender': gender or '',
                            'first_name': first_name or '',
                            'last_name': last_name or '',
                            'email': email or '',
                        }
                    )
                    if created:
                        self.stdout.write(self.style.SUCCESS(f'  Created fencer profile for: {identifier}'))
                    else:
                        # Update existing profile
                        if club:
                            fencer_profile.club = club
                        if phone:
                            fencer_profile.phone = phone
                        if gender:
                            fencer_profile.gender = gender
                        fencer_profile.save()
                        self.stdout.write(f'  Updated fencer profile for: {identifier}')
                else:
                    # Create profile without user
                    fencer_profile, created = FencerProfile.objects.get_or_create(
                        first_name=first_name,
                        last_name=last_name,
                        defaults={
                            'club': club,
                            'phone': phone or '',
                            'gender': gender or '',
                            'email': email or '',
                        }
                    )
                    if created:
                        self.stdout.write(self.style.SUCCESS(f'  Created fencer profile (no user) for: {identifier}'))
            else:
                self.stdout.write(f'  [DRY RUN] Would get/create fencer profile for: {identifier}')
                fencer_profile = type('FencerProfile', (), {'id': None})()  # Mock profile
            
            fencer_mapping[identifier] = fencer_profile
        
        return fencer_mapping

    def _import_events(self, events_data, dry_run):
        """Import events and return mapping of title+date to Event object."""
        event_mapping = {}
        
        for event_data in events_data:
            title = event_data.get('title')
            date_value = event_data.get('date')
            location = event_data.get('location', '')
            description = event_data.get('description', '')
            event_type = event_data.get('event_type', 'other')
            gender = event_data.get('gender', 'V')
            participants_count = event_data.get('participants_count')
            external_link = event_data.get('external_link', '')
            
            # Parse date
            from datetime import date as date_type
            if isinstance(date_value, datetime):
                event_date = date_value.date()
            elif isinstance(date_value, date_type):
                event_date = date_value
            elif isinstance(date_value, str):
                try:
                    # Try parsing as date string
                    event_date = datetime.strptime(date_value, '%Y-%m-%d').date()
                except:
                    try:
                        # Try parsing datetime and extract date
                        event_date = datetime.strptime(date_value, '%Y-%m-%d %H:%M:%S').date()
                    except:
                        self.stdout.write(self.style.ERROR(f'  Invalid date format: {date_value}'))
                        continue
            else:
                self.stdout.write(self.style.ERROR(f'  Invalid date type: {date_value}'))
                continue
            
            # Normalize event_type
            event_type_map = {
                'tournament': 'tournament',
                'turnaj': 'tournament',
                'humanitarian': 'humanitarian',
                'humanitární': 'humanitarian',
                'other': 'other',
                'ostatní': 'other',
            }
            event_type = event_type_map.get(event_type.lower(), 'other')
            
            # Normalize gender
            gender_map = {'M': 'M', 'Z': 'Z', 'Ž': 'Z', 'V': 'V', 'Vše': 'V', 'All': 'V'}
            gender = gender_map.get(str(gender).strip(), 'V')
            
            # Parse participants_count
            participants_count_value = None
            if participants_count:
                try:
                    participants_count_value = int(participants_count)
                except (ValueError, TypeError):
                    pass
            
            # Create event
            if not dry_run:
                event, created = Event.objects.get_or_create(
                    title=title,
                    date=event_date,
                    defaults={
                        'location': location or '',
                        'description': description or '',
                        'event_type': event_type,
                        'gender': gender,
                        'participants_count': participants_count_value,
                        'external_link': external_link or '',
                    }
                )
                if created:
                    self.stdout.write(self.style.SUCCESS(f'  Created event: {title} ({event_date})'))
                else:
                    self.stdout.write(f'  Event already exists: {title} ({event_date})')
            else:
                self.stdout.write(f'  [DRY RUN] Would get/create event: {title} ({event_date})')
                event = type('Event', (), {'id': None, 'title': title})()  # Mock event
            
            # Store mapping using title and date as key
            key = f"{title}|{event_date}"
            event_mapping[key] = event
            # Also allow lookup by title only
            if title not in event_mapping:
                event_mapping[title] = event
        
        return event_mapping

    def _import_participations(self, participations_data, fencer_mapping, event_mapping, dry_run):
        """Import event participations."""
        created_count = 0
        skipped_count = 0
        
        for participation_data in participations_data:
            fencer_identifier = participation_data.get('fencer_identifier')
            event_title = participation_data.get('event_title')
            date_value = participation_data.get('date')
            position = participation_data.get('position')
            wins = participation_data.get('wins', 0) or 0
            losses = participation_data.get('losses', 0) or 0
            touches_scored = participation_data.get('touches_scored', 0) or 0
            touches_received = participation_data.get('touches_received', 0) or 0
            
            # Find fencer
            fencer = fencer_mapping.get(fencer_identifier)
            if not fencer:
                self.stdout.write(self.style.WARNING(
                    f'  Fencer not found: {fencer_identifier}'
                ))
                skipped_count += 1
                continue
            
            # Find event
            event = None
            if date_value:
                # Try to parse date to match with event
                try:
                    from datetime import date as date_type
                    if isinstance(date_value, datetime):
                        date_key = date_value.date()
                    elif isinstance(date_value, date_type):
                        date_key = date_value
                    elif isinstance(date_value, str):
                        date_key = datetime.strptime(date_value, '%Y-%m-%d').date()
                    else:
                        date_key = None
                    
                    if date_key and event_title:
                        key = f"{event_title}|{date_key}"
                        event = event_mapping.get(key)
                except:
                    pass
            
            if not event and event_title:
                event = event_mapping.get(event_title)
            
            if not event:
                self.stdout.write(self.style.WARNING(
                    f'  Event not found: {event_title} (date: {date_value})'
                ))
                skipped_count += 1
                continue
            
            # Create participation
            if not dry_run:
                if hasattr(fencer, 'id') and hasattr(event, 'id'):  # Real objects
                    participation, created = EventParticipation.objects.get_or_create(
                        fencer=fencer,
                        event=event,
                        defaults={
                            'position': int(position) if position else None,
                            'wins': int(wins) if wins else 0,
                            'losses': int(losses) if losses else 0,
                            'touches_scored': int(touches_scored) if touches_scored else 0,
                            'touches_received': int(touches_received) if touches_received else 0,
                        }
                    )
                    if created:
                        created_count += 1
                        self.stdout.write(self.style.SUCCESS(
                            f'  Created participation: {fencer_identifier} -> {event.title}'
                        ))
                    else:
                        self.stdout.write(f'  Participation already exists: {fencer_identifier} -> {event.title}')
                else:
                    self.stdout.write(self.style.WARNING('  Cannot create participation in dry run mode'))
            else:
                self.stdout.write(f'  [DRY RUN] Would create participation: {fencer_identifier} -> {event_title}')
                created_count += 1
        
        self.stdout.write(self.style.SUCCESS(
            f'\nParticipations: {created_count} created, {skipped_count} skipped'
        ))


